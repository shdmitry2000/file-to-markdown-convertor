"""openpyxl -> :mod:`model`.

An .xlsx stores a formula cell twice: the formula and the result Excel cached at
its last save. openpyxl surfaces one or the other depending on ``data_only``, so
the workbook is opened twice and the two views are zipped together.

Everything else read here is metadata Excel already carries — merged ranges,
Tables, freeze panes, autofilter, defined names, comments, number formats,
indent/outline levels. Those are what let labels be derived deterministically
instead of guessed by a model.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from .model import Cell, Rect, Sheet, TableDef, WorkbookModel

logger = logging.getLogger(__name__)

DEFAULT_MAX_CELLS = 200_000


def _rect(ref: str) -> Rect:
    left, top, right, bottom = range_boundaries(ref)
    return Rect(top=top, left=left, bottom=bottom, right=right)


def _formula_text(value: Any) -> tuple[str | None, bool]:
    """Return (formula without '=', is_array). Non-formulas give (None, False)."""
    if isinstance(value, ArrayFormula):
        return (value.text or "").lstrip("="), True
    if isinstance(value, str) and value.startswith("="):
        return value[1:], False
    return None, False


def _freeze(sheet: Sheet, worksheet: Worksheet) -> None:
    panes = worksheet.freeze_panes
    if not panes:
        return
    try:
        row, col = coordinate_to_tuple(panes)
    except (ValueError, TypeError):
        logger.debug("unparsable freeze_panes %r on %s", panes, worksheet.title)
        return
    # freeze_panes names the first UNfrozen cell: 'A3' freezes rows 1-2.
    sheet.freeze_row = max(0, row - 1)
    sheet.freeze_col = max(0, col - 1)


def _tables(worksheet: Worksheet) -> list[TableDef]:
    tables = []
    for table in worksheet.tables.values():
        try:
            rect = _rect(table.ref)
        except (ValueError, TypeError):
            logger.debug("unparsable table ref %r on %s", table.ref, worksheet.title)
            continue
        tables.append(
            TableDef(
                name=table.displayName or table.name or "",
                rect=rect,
                header_row_count=table.headerRowCount or 0,
                columns=[column.name for column in (table.tableColumns or [])],
            )
        )
    return tables


def _defined_names(workbook) -> dict[str, str]:
    names: dict[str, str] = {}
    for name, definition in workbook.defined_names.items():
        names[name] = definition.attr_text or ""
    # Sheet-local names shadow nothing but still label cells; qualify to keep
    # them distinct from workbook-level names of the same text.
    for worksheet in workbook.worksheets:
        for name, definition in getattr(worksheet, "defined_names", {}).items():
            names.setdefault(f"{worksheet.title}!{name}", definition.attr_text or "")
    return names


def read_workbook(path: str | Path, max_cells: int = DEFAULT_MAX_CELLS) -> WorkbookModel:
    """Load *path* into a :class:`WorkbookModel`.

    Args:
        path: .xlsx/.xlsm file. Legacy .xls is not supported by openpyxl.
        max_cells: budget across the whole workbook. Sheets that hit it are
            marked ``truncated`` — the renderer says so rather than silently
            emitting a short table.

    Note:
        Read-only mode is NOT used: it drops comments, tables and styles, which
        are the highest-confidence label sources.
    """
    path = Path(path)
    formulas_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)

    model = WorkbookModel(defined_names=_defined_names(formulas_wb), source_name=path.name)
    budget = max_cells

    for index, worksheet in enumerate(formulas_wb.worksheets):
        values_ws = values_wb[worksheet.title]
        sheet = Sheet(
            name=worksheet.title,
            index=index,
            rtl=bool(worksheet.sheet_view.rightToLeft),
            hidden=worksheet.sheet_state != "visible",
            merged=[_rect(str(rng)) for rng in worksheet.merged_cells.ranges],
            tables=_tables(worksheet),
            hidden_rows={
                row for row, dim in worksheet.row_dimensions.items() if dim.hidden
            },
            hidden_cols={
                dim.min
                for dim in worksheet.column_dimensions.values()
                if dim.hidden and dim.min
            },
        )
        _freeze(sheet, worksheet)
        if worksheet.auto_filter and worksheet.auto_filter.ref:
            try:
                sheet.autofilter = _rect(worksheet.auto_filter.ref)
            except (ValueError, TypeError):
                logger.debug("unparsable autofilter on %s", worksheet.title)

        for row in worksheet.iter_rows():
            if budget <= 0:
                sheet.truncated = True
                logger.warning(
                    "excel: cell budget (%d) exhausted on sheet %r — output truncated",
                    max_cells,
                    worksheet.title,
                )
                break
            for source in row:
                formula, is_array = _formula_text(source.value)
                cached = values_ws.cell(row=source.row, column=source.column).value
                value = cached if formula else source.value
                comment = source.comment.text.strip() if source.comment else None
                fill = source.fill
                filled = bool(getattr(fill, "patternType", None))
                alignment = source.alignment
                if value is None and formula is None and comment is None:
                    continue
                budget -= 1
                sheet.cells[(source.row, source.column)] = Cell(
                    row=source.row,
                    col=source.column,
                    value=value,
                    formula=formula,
                    is_array_formula=is_array,
                    number_format=source.number_format or "General",
                    comment=comment,
                    bold=bool(source.font and source.font.bold),
                    filled=filled,
                    indent=int(alignment.indent or 0) if alignment else 0,
                    outline_level=int(
                        getattr(worksheet.row_dimensions[source.row], "outlineLevel", 0) or 0
                    ),
                )
                sheet.max_row = max(sheet.max_row, source.row)
                sheet.max_col = max(sheet.max_col, source.column)

        model.sheets.append(sheet)

    formulas_wb.close()
    values_wb.close()
    return model
