"""Generate the Excel fixture corpus for the excel converter.

Run (inside the container, openpyxl is already installed)::

    python tests/fixtures/excel/build_fixtures.py

Each workbook targets specific parts of the converter's scope; see README.md for
the mapping. Formula results are injected afterwards (see xlsx_cache.py) so the
files behave like Excel-saved workbooks, except where a fixture deliberately
omits them to exercise the "no cached value" path.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table

from xlsx_cache import inject_cached

OUT = Path(__file__).parent
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
NIS = '#,##0 [$₪-he-IL]'
PCT = "0.0%"


def _header(ws, row: int, first_col: int, labels: list[str]) -> None:
    for offset, label in enumerate(labels):
        cell = ws.cell(row=row, column=first_col + offset, value=label)
        cell.font = BOLD
        cell.fill = HEADER_FILL


# ---------------------------------------------------------------------------
# 1. Hebrew RTL loan book — the shape the partner's algorithm targets.
# ---------------------------------------------------------------------------
def build_heb_loans() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "הלוואות"
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "תיק הלוואות — רבעון 2, 2026"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    _header(ws, 2, 1, ["בנק", "סוג הלוואה", "יתרה", "ריבית", "תשלום חודשי"])
    loans = [
        ("הפועלים", "משכנתא", 850_000, 0.032),
        ("הפועלים", "צרכנית", 120_000, 0.079),
        ("לאומי", "משכנתא", 1_240_000, 0.041),
        ("לאומי", "עסקית", 430_000, 0.055),
        ("דיסקונט", "צרכנית", 96_000, 0.083),
    ]
    for index, (bank, kind, balance, rate) in enumerate(loans):
        row = 3 + index
        ws.cell(row=row, column=1, value=bank)
        ws.cell(row=row, column=2, value=kind)
        ws.cell(row=row, column=3, value=balance).number_format = NIS
        ws.cell(row=row, column=4, value=rate).number_format = PCT
        # Monthly payment — a per-row formula, so labels must resolve per row.
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}/12").number_format = NIS

    total_row = 3 + len(loans)
    ws.cell(row=total_row, column=1, value='סה"כ').font = BOLD
    ws.cell(row=total_row, column=3, value=f"=SUM(C3:C{total_row - 1})").number_format = NIS
    ws.cell(row=total_row, column=4, value=f"=C{total_row}/COUNT(C3:C{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{total_row - 1})").number_format = NIS

    # Author-supplied semantics: rungs L0 (comment) and L1 (defined name).
    ws.cell(row=total_row, column=3).comment = Comment(
        "סך יתרות ההלוואות בכל הבנקים", "מחלקת אשראי"
    )
    wb.defined_names.add(DefinedName("סך_יתרות", attr_text=f"הלוואות!$C${total_row}"))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{total_row - 1}"

    balances = [loan[2] for loan in loans]
    payments = [round(b * r / 12, 10) for _, _, b, r in loans]
    cache = {f"E{3 + i}": payments[i] for i in range(len(loans))}
    cache[f"C{total_row}"] = sum(balances)
    cache[f"D{total_row}"] = sum(balances) / len(balances)
    cache[f"E{total_row}"] = sum(payments)

    path = OUT / "heb_loans_rtl.xlsx"
    wb.save(path)
    inject_cached(path, {"הלוואות": cache})


# ---------------------------------------------------------------------------
# 2. Three-statement model — cross-sheet links and indented row hierarchy.
# ---------------------------------------------------------------------------
def build_three_statement() -> None:
    wb = Workbook()
    pnl = wb.active
    pnl.title = "רווח והפסד"
    balance = wb.create_sheet("מאזן")
    flow = wb.create_sheet("תזרים")
    for sheet in (pnl, balance, flow):
        sheet.sheet_view.rightToLeft = True

    years = [2024, 2025, 2026]
    _header(pnl, 1, 1, ["סעיף", *years])
    revenue = [12_000_000, 14_500_000, 17_200_000]
    cogs = [7_200_000, 8_400_000, 9_800_000]
    opex = [2_900_000, 3_300_000, 3_900_000]
    tax_rate = 0.23

    rows = {
        "הכנסות": revenue,
        "עלות המכר": cogs,
        "רווח גולמי": None,
        "הוצאות תפעול": opex,
        "רווח תפעולי": None,
        "מס": None,
        "רווח נקי": None,
    }
    indented = {"עלות המכר", "הוצאות תפעול", "מס"}
    for index, (label, values) in enumerate(rows.items()):
        row = 2 + index
        cell = pnl.cell(row=row, column=1, value=label)
        if label in indented:
            cell.alignment = Alignment(indent=2)
        else:
            cell.font = BOLD
        if values is not None:
            for col, value in enumerate(values, start=2):
                pnl.cell(row=row, column=col, value=value).number_format = NIS

    gross = [r - c for r, c in zip(revenue, cogs)]
    operating = [g - o for g, o in zip(gross, opex)]
    tax = [round(o * tax_rate, 10) for o in operating]
    net = [o - t for o, t in zip(operating, tax)]
    pnl_cache: dict[str, object] = {}
    for col_index, letter in enumerate("BCD"):
        pnl[f"{letter}4"] = f"={letter}2-{letter}3"          # רווח גולמי
        pnl[f"{letter}6"] = f"={letter}4-{letter}5"          # רווח תפעולי
        pnl[f"{letter}7"] = f"={letter}6*{tax_rate}"         # מס
        pnl[f"{letter}8"] = f"={letter}6-{letter}7"          # רווח נקי
        pnl_cache |= {
            f"{letter}4": gross[col_index],
            f"{letter}6": operating[col_index],
            f"{letter}7": tax[col_index],
            f"{letter}8": net[col_index],
        }

    _header(balance, 1, 1, ["סעיף", *years])
    cash = [1_100_000, 1_450_000, 1_900_000]
    receivables = [2_300_000, 2_700_000, 3_100_000]
    fixed_assets = [5_000_000, 5_400_000, 6_100_000]
    for row, (label, values) in enumerate(
        {"מזומן": cash, "לקוחות": receivables, "רכוש קבוע": fixed_assets}.items(), start=2
    ):
        cell = balance.cell(row=row, column=1, value=label)
        cell.alignment = Alignment(indent=2)
        for col, value in enumerate(values, start=2):
            balance.cell(row=row, column=col, value=value).number_format = NIS
    balance.cell(row=5, column=1, value='סה"כ נכסים').font = BOLD
    balance.cell(row=6, column=1, value="עודפים (מרווח נקי)").font = BOLD
    balance_cache: dict[str, object] = {}
    for col_index, letter in enumerate("BCD"):
        balance[f"{letter}5"] = f"=SUM({letter}2:{letter}4)"
        balance[f"{letter}6"] = f"='רווח והפסד'!{letter}8"   # cross-sheet link
        balance_cache[f"{letter}5"] = cash[col_index] + receivables[col_index] + fixed_assets[col_index]
        balance_cache[f"{letter}6"] = net[col_index]

    _header(flow, 1, 1, ["סעיף", *years])
    depreciation = [400_000, 460_000, 520_000]
    flow.cell(row=2, column=1, value="רווח נקי").font = BOLD
    flow.cell(row=3, column=1, value="פחת").alignment = Alignment(indent=2)
    flow.cell(row=4, column=1, value="תזרים מפעילות שוטפת").font = BOLD
    flow_cache: dict[str, object] = {}
    for col_index, letter in enumerate("BCD"):
        flow[f"{letter}2"] = f"='רווח והפסד'!{letter}8"
        flow.cell(row=3, column=2 + col_index, value=depreciation[col_index]).number_format = NIS
        flow[f"{letter}4"] = f"={letter}2+{letter}3"
        flow_cache[f"{letter}2"] = net[col_index]
        flow_cache[f"{letter}4"] = net[col_index] + depreciation[col_index]

    path = OUT / "three_statement_heb.xlsx"
    wb.save(path)
    inject_cached(
        path,
        {"רווח והפסד": pnl_cache, "מאזן": balance_cache, "תזרים": flow_cache},
    )


# ---------------------------------------------------------------------------
# 3. Several independent tables on one sheet — region detection.
# ---------------------------------------------------------------------------
def build_multi_region() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Q1 Revenue by Region"
    ws["A1"].font = BOLD
    _header(ws, 2, 1, ["Region", "Revenue"])
    for row, (region, value) in enumerate(
        [("North", 320_000), ("South", 280_000), ("Center", 510_000)], start=3
    ):
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=value)
    ws["B6"] = "=SUM(B3:B5)"
    ws["A6"] = "Total"

    # Second table, side by side — separated by an empty column D.
    ws["E1"] = "Headcount by Department"
    ws["E1"].font = BOLD
    _header(ws, 2, 5, ["Department", "People"])
    for row, (dept, value) in enumerate(
        [("R&D", 42), ("Sales", 18), ("Ops", 9)], start=3
    ):
        ws.cell(row=row, column=5, value=dept)
        ws.cell(row=row, column=6, value=value)
    ws["E6"] = "Total"
    ws["F6"] = "=SUM(F3:F5)"

    # Third table, below both — separated by empty rows 7-8.
    ws["A9"] = "Cost per Employee"
    ws["A9"].font = BOLD
    _header(ws, 10, 1, ["Metric", "Value"])
    ws["A11"] = "Revenue per head"
    ws["B11"] = "=B6/F6"

    path = OUT / "multi_region.xlsx"
    wb.save(path)
    inject_cached(
        path,
        {"Summary": {"B6": 1_110_000, "F6": 69, "B11": 1_110_000 / 69}},
    )


# ---------------------------------------------------------------------------
# 4. Cross-tab with a merged two-level header and totals on both axes.
# ---------------------------------------------------------------------------
def build_crosstab() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"

    ws["B1"] = 2026
    ws["B1"].font = BOLD
    ws.merge_cells("B1:E1")           # spanning header above the quarters
    _header(ws, 2, 2, ["Q1", "Q2", "Q3", "Q4"])
    ws["F2"] = "Total"
    ws["F2"].font = BOLD
    ws["A2"] = "Cost centre"
    ws["A2"].font = BOLD

    data = {
        "Marketing": [120, 135, 150, 160],
        "Engineering": [400, 420, 445, 470],
        "Support": [90, 95, 92, 101],
    }
    cache: dict[str, object] = {}
    for index, (name, quarters) in enumerate(data.items()):
        row = 3 + index
        ws.cell(row=row, column=1, value=name)
        for col, value in enumerate(quarters, start=2):
            ws.cell(row=row, column=col, value=value)
        ws.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
        cache[f"F{row}"] = sum(quarters)

    total_row = 3 + len(data)
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    for col_letter in "BCDEF":
        ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
    columns = list(zip(*data.values()))
    for index, letter in enumerate("BCDE"):
        cache[f"{letter}{total_row}"] = sum(columns[index])
    cache[f"F{total_row}"] = sum(sum(q) for q in data.values())

    path = OUT / "crosstab_multiheader.xlsx"
    wb.save(path)
    inject_cached(path, {"Budget": cache})


# ---------------------------------------------------------------------------
# 5. Reference dialects and failure modes.
# ---------------------------------------------------------------------------
def build_edge_cases() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Edge"

    _header(ws, 1, 1, ["item", "qty", "price"])
    for row, (item, qty, price) in enumerate(
        [("bolt", 100, 1.5), ("nut", 250, 0.4), ("washer", 80, 0.25)], start=2
    ):
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=price)
    ws.add_table(Table(displayName="Parts", ref="A1:C4"))

    ws["E1"] = "array"
    ws["F1"] = ArrayFormula("F1", "=SUM(B2:B4*C2:C4)")      # array formula
    ws["E2"] = "whole column"
    ws["F2"] = "=SUM(B:B)"                                   # whole-column ref
    ws["E3"] = "structured"
    ws["F3"] = "=SUM(Parts[qty])"                            # structured table ref
    ws["E4"] = "indirect"
    ws["F4"] = '=SUM(INDIRECT("B2:B4"))'                     # dynamic, unresolvable
    ws["E5"] = "error"
    ws["F5"] = "=B2/0"                                       # cached as an error value
    ws["E6"] = "cross sheet"
    ws["F6"] = "='Notes Sheet'!B2"                           # quoted sheet name
    ws["E7"] = "hidden row below"
    ws.row_dimensions[8].hidden = True
    ws["A8"] = "hidden item"
    ws["B8"] = 999

    notes = wb.create_sheet("Notes Sheet")
    notes["A2"] = "answer"
    notes["B2"] = 42

    # Deliberately NOT cached — exercises the "formula with no result" path.
    uncached = wb.create_sheet("NoCache")
    uncached["A1"] = 10
    uncached["A2"] = 20
    uncached["A3"] = "=SUM(A1:A2)"

    path = OUT / "edge_cases.xlsx"
    wb.save(path)
    inject_cached(
        path,
        {
            "Edge": {
                "F1": 100 * 1.5 + 250 * 0.4 + 80 * 0.25,
                "F2": 100 + 250 + 80 + 999,
                "F3": 100 + 250 + 80,
                "F4": 100 + 250 + 80,
                "F5": "#DIV/0!",
                "F6": 42,
            }
        },
    )


def main() -> None:
    for build in (
        build_heb_loans,
        build_three_statement,
        build_multi_region,
        build_crosstab,
        build_edge_cases,
    ):
        build()
        print(f"built {build.__name__.removeprefix('build_')}")


if __name__ == "__main__":
    main()
